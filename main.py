#!/usr/bin/env python


import os
import argparse
import json
import asyncio
import schedule
from typing import Tuple, List

from datetime import datetime
from contextlib import suppress
from asgiref.sync import sync_to_async
from numpy import ndarray

import pyfiglet
from openpyxl import Workbook, load_workbook
from termcolor import colored

from selenium import webdriver
from selenium.webdriver.chrome.service import Service

import pickle5 as pickle
from sklearn.ensemble import RandomForestRegressor



class Predictor:
    def __init__(
                    self,
                    data_set:int,
                    update_time:int,
                    training:bool
                ):
        self.data_file:str = 'hash.xlsx'
        self.path:str = './data'
        self.path_complete:str = self.path + '/' + self.data_file

        self.workbook:Workbook = self.get_workbook()
        self.webdriver:str = './webdriver/chromedriver'

        self.update_time:int = update_time
        self.data_set:int = data_set
        self.training:int = training

        self.actual_game:dict = dict()
        self.last_results:list = list()
        self.all_values:list = self.get_all_values()

        self.ai = self.AI(
                            all_values=self.all_values,
                            last_results=self.last_results,
                            data_set=self.data_set
                        )

        schedule.every(self.update_time).minutes.do(self.update_data)


    class AI:
        def __init__(
                        self, 
                        all_values:list, 
                        last_results:list,
                        data_set=int
                    ):
            self.data:list = all_values
            self.last_results:list = last_results
            self.data_set:int = data_set

            self.model_path:str = './models'
            self.regression_model_path:str = self.model_path + '/' + 'regression_model.pkl'
            self.regressor_params:dict = {'n_estimators': 400}

        
        def normalize_data(self, value:int) -> int:
            if value <= 199: return 1
            elif value >= 200 and value <= 299: return 2
            elif value >= 300 and value <= 399: return 3
            elif value >= 400 and value <= 499: return 4
            elif value >= 500 and value <= 599: return 5
            elif value >= 600 and value <= 699: return 6
            elif value >= 700 and value <= 799: return 7
            elif value >= 800 and value <= 899: return 8
            elif value >= 900 and value <= 999: return 9
            elif value >= 1000: return 10
            
        
        def get_input(self) -> list:
            return [[self.normalize_data(i) for i in self.last_results]]


        def format_value(self, value:int) -> str:
            if value >= 101: value = '{:.2f}'.format(round(int(value)/100, 3))
            else: value = 0
            return f'{value}x'


        async def get_X_and_y(self, normalize:bool=True) -> Tuple[List, List]:
            if normalize:
                data = list()
                for i in range(len(self.data)-self.data_set):
                    sets = [self.normalize_data(self.data[i])]
                    for j in range(1, self.data_set):
                        sets.append(self.normalize_data(self.data[i+j]))
                    data.append(sets)
            else:
                data = [self.data[i:i+self.data_set] for i in range(len(self.data)-self.data_set)]
            X, y = data[:-1], data[1:]
            return X, y
        

        async def history(self) -> None:
            data = list()
            X, y = await self.get_X_and_y(normalize=False)
            for i in range(len(X)-self.data_set):
                j = X[i]
                if all(j[k] <= self.last_results[k]+50 and j[k]+80 >= self.last_results[k] for k in range(self.data_set)):
                    value:int = X[i+self.data_set][0]
                    if value not in data:
                        data.append(value)
            if len(data) >= 1:
                data.reverse()
                if len(data) >= 6:
                    data = data[:5]
                data = ', '.join(str(self.format_value(i)) for i in data)
                message:str = 'Crash anterior en una circunstancia similar'
                print(f"{message}: {colored(f'[ {data} ]', 'light_grey')}")


        async def prediction(self) -> None:
            if not os.path.exists(self.regression_model_path):
                X, y = await self.get_X_and_y()
                model = await sync_to_async(RandomForestRegressor(**self.regressor_params).fit)(X=X, y=y)
                with open(self.regression_model_path, 'wb') as f:
                    pickle.dump(model, f)
            else:
                model:RandomForestRegressor = await sync_to_async(pickle.load)(open(self.regression_model_path, 'rb'))
            predict:ndarray = await sync_to_async(model.predict)(self.get_input())
            result = predict[0][-1]
            result:float = round(result,1)-1
            if result <= 2.0: message = colored("Quizás no pase de 1x ⛔", "red")
            else: message = colored(f"En un rango de {result}x 🚀", "green")
            print(f'Estimado: {message}')
        

        async def upload_model(self) -> None:
            with open(self.regression_model_path, 'rb') as f:
                model = pickle.load(f)
            X, y = await self.get_X_and_y()
            model:RandomForestRegressor = await sync_to_async(RandomForestRegressor(**self.regressor_params).fit)(X=X, y=y)
            accuracy:float = await sync_to_async(model.score)(X, y)
            with open(self.regression_model_path, 'wb') as f:
                model = pickle.dump(model, f)
            print('Modelo de Regresion actualizado : [ {}% ]'.format(round(accuracy*100, 2)))
        
        
        async def run(self, task:str=None) -> None:
            if task is None:
                tasks = [
                            asyncio.create_task(self.prediction()),
                            asyncio.create_task(self.history())
                        ]
            elif task =='update':
                tasks = [
                            asyncio.create_task(self.upload_model())
                        ]
            await asyncio.gather(*tasks)
            print('\n')
    

    def get_workbook(self) -> Workbook:
        if not os.path.exists(self.path_complete):
            if not os.path.exists(self.path):
                os.makedirs(self.path)
            workbook = Workbook()
            sheet = workbook.active
            sheet['A1'].value = 'Date'
            sheet['B1'].value = 'Game ID'
            sheet['C1'].value = 'Hash'
            sheet['D1'].value = 'Crash'
        else:
            workbook = load_workbook(self.path_complete)
        return workbook
    

    def get_all_values(self) -> list:
        return [int(row[3]) for row in self.workbook['Sheet'].iter_rows(values_only=True) if str(row[3]).isdigit()]


    def update_data(self) -> None:
        print('Actualizando datos...')
        old_data:int = len(self.all_values)
        self.all_values:list = self.get_all_values()
        new_data:int = len(self.all_values)
        total_data:int = new_data - old_data
        print('{}: {}\n'.format(colored('Datos actualizados', 'green', attrs=["bold"]), total_data))
        asyncio.run(self.ai.run(task='update'))


    def save_data(self) -> None: 
        sheet = self.workbook.active
        sheet.append((self.actual_game['Date'], self.actual_game['Game ID'], self.actual_game['Hash'], self.actual_game['Crash']))
        self.workbook.save(self.path_complete)


    def start(self) -> None:
        print(pyfiglet.figlet_format('BitBust - AI', font='starwars', width=500))
        print(' '*15 + 'by github.io/luiserb  🚀\n')
        print('\n{}\n'.format(colored('Iniciando asistente de apuestas...', 'green', attrs=["bold"])))
        print('Registros disponibles: {}  [ {} MB ] \n'.format(colored("{}".format(len(self.all_values)), "green"), round(int(os.path.getsize(self.path_complete))/1000000, 2)))
        service = Service(self.webdriver)
        desired_capabilities = webdriver.DesiredCapabilities.CHROME.copy()
        desired_capabilities['goog:loggingPrefs'] = {'performance': 'ALL'}
        browser = webdriver.Chrome(service=service, desired_capabilities=desired_capabilities)
        browser.get('https://www.solcasino.io/crash')
        print('{}\n'.format(colored('Cargando websocket...', 'green')))
        if self.training: print('{}\n'.format(colored('Tomando data...', 'green')))
        while True:
            for data in browser.get_log('performance'):
                ws_data:dict = json.loads(data['message'])
                if ws_data['message']['method'] == 'Network.webSocketFrameReceived':
                    message:str = ws_data['message']['params']['response']['payloadData']
                    if '42["game_starting' in message or '42["game_crash' in message:
                        with suppress(Exception):
                            data:dict = json.loads(str(message.replace('42', '')))
                            if data[0] == 'game_starting':
                                self.actual_game['Game ID'] = data[1]['game_id']
                                self.actual_game['Date'] = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
                            elif data[0] == 'game_crash':
                                self.actual_game['Hash'] = data[1]['hash']
                                self.actual_game['Crash'] = data[1]['game_crash']
                                if self.actual_game.get('Game ID', None) is not None:
                                    self.save_data()
                                    self.last_results.append(self.actual_game['Crash'])
                                    print(f"Ultimo crash: {self.ai.format_value(self.last_results[-1])} 🚀")
                                    if self.training is False:
                                        if len(self.last_results) >= self.data_set:
                                            if len(self.last_results) >= self.data_set + 1:
                                                self.last_results.remove(self.last_results[0])
                                            asyncio.run(self.ai.run())
                                        else:
                                            print(f"Detectando patrones ({len(self.last_results)}/{self.data_set}) \n")
                                    else:
                                        print('\nResultados agregados: {}'.format(len(self.last_results)))
            if self.training is False: schedule.run_pending()



if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('-d', '--dataset', type=int, help='Número de conjunto de datos a escoger', default=8)
    parser.add_argument('-ut','--update_time', type=int, help='Tiempo de actualización de datos', default=18)
    parser.add_argument('-t', '--training', action='store_true',help='Obtención de datos')
    args = parser.parse_args()
    predictor = Predictor(
                            data_set=args.dataset, 
                            update_time=args.update_time,
                            training=args.training
                        )
    predictor.start()
